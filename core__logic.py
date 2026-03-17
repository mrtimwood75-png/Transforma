import re
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_DELIVERY_TYPE = "premium delivery"
DEFAULT_LOCATION = "bcwh"


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def files_dir() -> Path:
    return app_dir() / "files"


def default_template_path():
    folder = files_dir()
    if not folder.exists():
        return None

    matches = sorted(folder.glob("*.xlsx"))
    return matches[0] if matches else None


def read_text_file_bytes(file_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin1", "utf-16"):
        try:
            return file_bytes.decode(enc, errors="ignore")
        except Exception:
            pass
    return file_bytes.decode(errors="ignore")


def split_rows(text: str):
    rows = []
    for line in text.splitlines():
        rows.append([c.strip() for c in line.rstrip("\n\r").split("\t")])
    return rows


def flatten_non_empty(cells):
    return [str(c).strip() for c in cells if str(c).strip()]


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def regex_find(text: str, pattern: str, flags=0, default=""):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else default


def title_case_city(text: str) -> str:
    if not text:
        return ""
    return " ".join(w.capitalize() if not w.isupper() else w for w in text.split())


def normalise_phone(phone: str) -> str:
    return re.sub(r"[^\d+]", "", phone or "")


def extract_sales_order(text: str) -> str:
    m = re.search(r"\b(OS-\d{4,})\b", text, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return ""


def extract_notes(text: str) -> str:
    lines = [clean_spaces(x) for x in text.splitlines()]
    ignore = {
        "",
        "BoConcept signature",
        "Customer signature",
        "Tax code Amount origin Tax amount",
    }
    candidates = []
    for line in lines:
        if line in ignore:
            continue
        if "Terms & Conditions" in line:
            continue
        if "Payment of any deposit" in line:
            continue
        if "www.bit.ly/BoConceptTerms" in line:
            continue
        if "Trading as BoConcept" in line:
            continue
        if re.search(r"\bSales balance\b|\bBalance due\b|\bPrepayment\b", line, re.I):
            continue
        if any(tag in line.upper() for tag in ["ETA", "DRAWING", "LOANER", "NO ACTION REQUIRED"]):
            candidates.append(line)
    return " | ".join(dict.fromkeys(candidates))


def extract_header_data(text: str, rows):
    first_line = flatten_non_empty(rows[0]) if rows else []
    second_line = flatten_non_empty(rows[1]) if len(rows) > 1 else []

    customer_name = first_line[0] if len(first_line) > 0 else ""
    ship_address = first_line[1] if len(first_line) > 1 else ""

    city = ""
    postcode = ""

    if second_line:
        line2 = " ".join(second_line)
        m = re.search(r"^(.*?)\s+(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+(\d{4})$", line2, re.I)
        if m:
            city = clean_spaces(m.group(1))
            postcode = m.group(2)

    if not city:
        city = regex_find(text, r"\n([A-Z][A-Z\s,]+)\s+(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+\d{4}", re.I)
        city = clean_spaces(city).rstrip(",")

    if not postcode:
        postcode = regex_find(text, r"\b(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+(\d{4})\b", re.I)

    email = regex_find(text, r"\bE-Mail\s+([^\s]+@[^\s]+)")
    phone = regex_find(text, r"\bPhone\s+([0-9+ ]{6,})")
    if not phone:
        phone = regex_find(text, r"\bCustomer number\s+([0-9+ ]{6,})")

    return {
        "sales_order_number": extract_sales_order(text),
        "customer_name": customer_name,
        "ship_address": ship_address,
        "ship_zip": postcode,
        "ship_city": title_case_city(city),
        "phone": normalise_phone(phone),
        "email": email,
        "notes": extract_notes(text),
    }


def parse_qty(token: str) -> int:
    token = str(token).strip().replace(",", ".")
    try:
        return int(round(float(token)))
    except Exception:
        return 1


def normalise_fraction_dimensions(text: str) -> str:
    repl = {
        "½": ".5",
        "¼": ".25",
        "¾": ".75",
        "⅓": ".33",
        "⅔": ".67",
        "⅛": ".125",
        "⅜": ".375",
        "⅝": ".625",
        "⅞": ".875",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    return text


def extract_dimensions(description: str) -> str:
    d = normalise_fraction_dimensions(description or "")

    patterns = [
        r"H\s*\d+(?:[./]\d+)?\s*[xX]\s*W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*D\s*(\d+(?:[./]\d+)?)\s*cm",
        r"W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*D\s*(\d+(?:[./]\d+)?)(?:\s*[xX]\s*H\s*\d+(?:[./]\d+)?)?\s*cm",
        r"W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*L\s*(\d+(?:[./]\d+)?)\s*cm",
        r"\b(\d{2,4}(?:[./]\d+)?)\s*[xX]\s*(\d{2,4}(?:[./]\d+)?)\s*cm\b",
    ]

    for pat in patterns:
        m = re.search(pat, d, re.I)
        if m:
            return f"{m.group(1)} x {m.group(2)}"

    return ""


def extract_full_dimensions(description: str) -> str:
    d = normalise_fraction_dimensions(description or "")

    patterns = [
        r"H\s*\d+(?:[./]\d+)?(?:/\d+(?:[./]\d+)?)?\s*[xX]\s*W\s*\d+(?:[./]\d+)?\s*[xX]\s*D\s*\d+(?:[./]\d+)?\s*cm",
        r"W\s*\d+(?:[./]\d+)?\s*[xX]\s*D\s*\d+(?:[./]\d+)?(?:\s*[xX]\s*H\s*\d+(?:[./]\d+)?)?\s*cm",
        r"W\s*\d+(?:[./]\d+)?\s*[xX]\s*L\s*\d+(?:[./]\d+)?\s*cm",
        r"\b\d{2,4}(?:[./]\d+)?\s*[xX]\s*\d{2,4}(?:[./]\d+)?\s*cm\b",
    ]

    for pat in patterns:
        m = re.search(pat, d, re.I)
        if m:
            return clean_spaces(m.group(0).replace(" ", ""))

    return ""


# ---------------- OLD FORMAT PARSER ----------------

def parse_items_from_ascii_text(text: str):
    lines = [line.rstrip() for line in text.splitlines()]

    item_start_re = re.compile(
        r"^\s*(\d{4,25})\s+(-?\d+,\d+)\s+(.+?)\s+(?:-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(?:-?\d{1,3}(?:\.\d{3})*,\d{2})\s*$"
    )

    item_start_loose = re.compile(
        r"^\s*(\d{4,25})\s+(-?\d+,\d+)\s+(.+)$"
    )

    stop_re = re.compile(
        r"^\s*(Total for |Tax code\b|gst\d|Sales balance\b|Prepayment\b|Balance due\b)",
        re.I,
    )

    skip_re = re.compile(
        r"^\s*(Article\s+Qty\.?\s+Description|Discount|in pct|per Unit|Price|Amount|Confirmation|Page\s+\d+\s+of\s+\d+)\b",
        re.I,
    )

    footer_re = re.compile(
        r"^\s*(BC Brisbane Pty Ltd|ABN |Tel: |BoConcept.? signature|Customer signature|This order is subject to|Payment of any deposit|You can view BoConcept)",
        re.I,
    )

    items = []
    current = None
    in_items = False

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            continue

        if "Article" in line and "Qty" in line and "Description" in line:
            in_items = True
            continue

        if not in_items:
            continue

        if stop_re.search(line):
            if current:
                current["description"] = clean_spaces(current["description"])
                current["volume"] = extract_dimensions(current["description"])
                current["dimensions"] = extract_full_dimensions(current["description"])
                items.append(current)
                current = None

            if re.match(r"^\s*Tax code\b", line, re.I):
                break

            continue

        if skip_re.search(line) or footer_re.search(line):
            continue

        m = item_start_re.match(line)
        if not m:
            m = item_start_loose.match(line)

        if m:
            if current:
                current["description"] = clean_spaces(current["description"])
                current["volume"] = extract_dimensions(current["description"])
                current["dimensions"] = extract_full_dimensions(current["description"])
                items.append(current)

            sku, qty_raw, desc = m.groups()
            current = {
                "sku": sku,
                "qty": parse_qty(qty_raw),
                "description": clean_spaces(desc),
                "volume": "",
                "dimensions": "",
            }
            continue

        if current:
            if (
                not line.startswith("Total for")
                and "Page " not in line
                and "Sales order " not in line
                and "BC Brisbane Pty Ltd" not in line
            ):
                current["description"] += " " + line

    if current:
        current["description"] = clean_spaces(current["description"])
        current["volume"] = extract_dimensions(current["description"])
        current["dimensions"] = extract_full_dimensions(current["description"])
        items.append(current)

    return items


def parse_old_format_order_bytes(file_bytes: bytes):
    text = read_text_file_bytes(file_bytes)
    rows = split_rows(text)
    header_data = extract_header_data(text, rows)
    items = parse_items_from_ascii_text(text)
    return [(header_data, items)]


# ---------------- NEW FORMAT PARSER ----------------

def split_lines(text: str):
    return [line.rstrip("\r") for line in text.splitlines()]


def extract_new_format_header_data(header_lines):
    header_text = "\n".join(header_lines)
    header_parts = []
    for line in header_lines:
        parts = [p.strip() for p in line.split("\t") if p.strip()]
        header_parts.extend(parts)

    sales_order = ""
    customer_name = ""
    ship_address = ""
    ship_city = ""
    ship_zip = ""
    phone = ""
    email = ""
    notes = ""

    for i, part in enumerate(header_parts[:-1]):
        if part.lower() == "sales order":
            sales_order = header_parts[i + 1].upper()
            break

    for i, part in enumerate(header_parts[:-1]):
        if part.lower() == "telephone":
            phone = normalise_phone(header_parts[i + 1])
            break

    if not sales_order:
        m = re.search(r"\b(OS-\d{4,})\b", header_text, re.I)
        if m:
            sales_order = m.group(1).upper()

    meaningful = []
    for p in header_parts:
        pl = p.lower()
        if pl in {
            "packinglist - order", "shop", "sales order", "shop order", "date", "page",
            "ic-deliverymode", "turn no.", "car", "week", "day", "recipient", "tour date",
            "customer account", "name", "australia", "telephone"
        }:
            continue
        if re.fullmatch(r"\d+/\d+/\d+", p):
            continue
        if re.fullmatch(r"tr-\d+", p, re.I):
            continue
        if re.fullmatch(r"os-\d+", p, re.I):
            continue
        if re.fullmatch(r"\d{8,}", p):
            continue
        if re.fullmatch(r"\d+", p):
            continue
        meaningful.append(p)

    if meaningful:
        customer_name = meaningful[0]

    address_candidates = []
    for p in meaningful[1:]:
        if re.search(r"\b(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\b\s+\d{4}", p, re.I):
            address_candidates.append(p)
        elif not ship_address:
            ship_address = p

    for p in address_candidates:
        m = re.search(r"^(.*?)(?:,)?\s+(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+(\d{4})$", p, re.I)
        if m:
            ship_city = title_case_city(clean_spaces(m.group(1)))
            ship_zip = m.group(2)
            break

    return {
        "sales_order_number": sales_order,
        "customer_name": customer_name,
        "ship_address": ship_address,
        "ship_zip": ship_zip,
        "ship_city": ship_city,
        "phone": phone,
        "email": email,
        "notes": notes,
    }


def extract_new_format_item_notes(extra_lines):
    cleaned = []
    for line in extra_lines:
        line = clean_spaces(line)
        if not line:
            continue
        if line.lower().startswith("total volume"):
            continue
        if line.lower().startswith("receipt"):
            continue
        cleaned.append(line)
    return " | ".join(dict.fromkeys(cleaned))


def parse_new_format_item_row(parts):
    if len(parts) < 6:
        return None

    item_number = parts[4].strip() if len(parts) > 4 else ""
    description = clean_spaces(parts[5]) if len(parts) > 5 else ""

    if not item_number or not description:
        return None

    qty = parse_qty(parts[2] if len(parts) > 2 else "1")

    volume_value = ""
    if len(parts) > 13:
        volume_value = parts[13].strip().replace(",", ".")

    return {
        "sku": item_number,
        "qty": qty,
        "description": description,
        "volume": volume_value,
        "dimensions": extract_full_dimensions(description),
    }


def parse_new_format_order_bytes(file_bytes: bytes):
    text = read_text_file_bytes(file_bytes)
    lines = split_lines(text)

    sections = []
    current_header = []
    current_items = []
    current_note_lines = []
    in_items = False

    for raw_line in lines:
        if not raw_line.strip():
            continue

        parts = [p.strip() for p in raw_line.split("\t")]

        if parts and parts[0] == "Location" and len(parts) > 5 and parts[4] == "Item number":
            in_items = True
            continue

        if parts and parts[0] == "Total volume":
            in_items = False
            continue

        if parts and parts[0] == "Receipt":
            if current_header or current_items:
                header_data = extract_new_format_header_data(current_header)
                note_text = extract_new_format_item_notes(current_note_lines)
                if note_text:
                    header_data["notes"] = note_text
                sections.append((header_data, current_items))
            current_header = []
            current_items = []
            current_note_lines = []
            in_items = False
            continue

        if not in_items and parts and re.fullmatch(r"\d{8,15}", parts[0]) and len(parts) > 1:
            if current_header or current_items:
                header_data = extract_new_format_header_data(current_header)
                note_text = extract_new_format_item_notes(current_note_lines)
                if note_text:
                    header_data["notes"] = note_text
                sections.append((header_data, current_items))
                current_items = []
                current_note_lines = []
            current_header = [raw_line]
            continue

        if in_items:
            item = parse_new_format_item_row(parts)
            if item:
                current_items.append(item)
            else:
                current_note_lines.append(raw_line)
        else:
            current_header.append(raw_line)

    if current_header or current_items:
        header_data = extract_new_format_header_data(current_header)
        note_text = extract_new_format_item_notes(current_note_lines)
        if note_text:
            header_data["notes"] = note_text
        sections.append((header_data, current_items))

    return sections


# ---------------- FORMAT DETECTION ----------------

def detect_report_format(text: str) -> str:
    text_upper = text.upper()

    old_score = 0
    new_score = 0

    if "ARTICLE" in text_upper and "QTY" in text_upper and "DESCRIPTION" in text_upper:
        old_score += 3
    if "TAX CODE" in text_upper:
        old_score += 1
    if "SALES BALANCE" in text_upper:
        old_score += 1
    if "PREPAYMENT" in text_upper or "BALANCE DUE" in text_upper:
        old_score += 1

    if "PACKINGLIST - ORDER" in text_upper:
        new_score += 3
    if "ITEM NUMBER" in text_upper and "PALLET ID" in text_upper:
        new_score += 2
    if "RECEIPT" in text_upper:
        new_score += 1
    if "TOTAL VOLUME" in text_upper:
        new_score += 1

    if new_score > old_score and new_score >= 3:
        return "new"

    if old_score >= 3:
        return "old"

    raise ValueError("Could not determine report format from uploaded file.")


def parse_order_bytes(file_bytes: bytes):
    text = read_text_file_bytes(file_bytes)
    report_format = detect_report_format(text)

    if report_format == "old":
        return parse_old_format_order_bytes(file_bytes)

    if report_format == "new":
        return parse_new_format_order_bytes(file_bytes)

    raise ValueError("Unsupported report format.")


# ---------------- COMMON EXPORT LOGIC ----------------

def map_headers(ws):
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip().lower()] = cell.column
    return header_map


def set_text(cell, value: str):
    cell.value = "" if value is None else str(value)
    cell.number_format = "@"


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        s = ws.cell(source_row, col)
        t = ws.cell(target_row, col)
        if s.has_style:
            t._style = s._style.copy()
        if s.font:
            t.font = s.font.copy()
        if s.fill:
            t.fill = s.fill.copy()
        if s.border:
            t.border = s.border.copy()
        if s.alignment:
            t.alignment = s.alignment.copy()
        if s.protection:
            t.protection = s.protection.copy()
        if s.number_format:
            t.number_format = s.number_format


def build_row_values(header_data: dict, item: dict):
    return {
        "sales order number": header_data.get("sales_order_number", ""),
        "delivery type": DEFAULT_DELIVERY_TYPE,
        "customer name": header_data.get("customer_name", ""),
        "sku number": item.get("sku", ""),
        "product description": item.get("description", ""),
        "quantity": item.get("qty", ""),
        "location": DEFAULT_LOCATION,
        "ship address": header_data.get("ship_address", ""),
        "ship zip": header_data.get("ship_zip", ""),
        "ship city": header_data.get("ship_city", ""),
        "phone": header_data.get("phone", ""),
        "weight": "",
        "volume(unit)": item.get("volume", ""),
        "emailaddress": header_data.get("email", ""),
        "notes": header_data.get("notes", ""),
        "dimensions": item.get("dimensions", ""),
    }


def fill_workbook_from_rows(template_bytes: bytes, all_rows: list) -> bytes:
    wb = load_workbook(BytesIO(template_bytes))
    ws = wb[wb.sheetnames[0]]
    header_map = map_headers(ws)
    max_col = ws.max_column

    required_headers = [
        "sales order number",
        "delivery type",
        "customer name",
        "sku number",
        "product description",
        "quantity",
        "location",
        "ship address",
        "ship zip",
        "ship city",
        "phone",
        "weight",
        "volume(unit)",
        "emailaddress",
        "notes",
    ]
    for h in required_headers:
        if h not in header_map:
            raise ValueError(f"Template header not found: {h}")

    has_dimensions_col = "dimensions" in header_map
    base_row = 2

    if not all_rows:
        all_rows = [{
            "sales order number": "",
            "delivery type": DEFAULT_DELIVERY_TYPE,
            "customer name": "",
            "sku number": "",
            "product description": "",
            "quantity": "",
            "location": DEFAULT_LOCATION,
            "ship address": "",
            "ship zip": "",
            "ship city": "",
            "phone": "",
            "weight": "",
            "volume(unit)": "",
            "emailaddress": "",
            "notes": "",
            "dimensions": "",
        }]

    if len(all_rows) > 1:
        ws.insert_rows(base_row + 1, amount=len(all_rows) - 1)
        for r in range(base_row + 1, base_row + len(all_rows)):
            copy_row_style(ws, base_row, r, max_col)

    for i, row_values in enumerate(all_rows, start=base_row):
        row_data = dict(row_values)
        if not has_dimensions_col:
            row_data.pop("dimensions", None)

        for header, col in header_map.items():
            if header not in row_data:
                continue

            cell = ws.cell(i, col)
            value = row_data[header]

            if header in {"sku number", "phone", "ship zip", "sales order number"}:
                set_text(cell, value)
            elif header == "quantity":
                cell.value = "" if value == "" else int(value)
            else:
                cell.value = value

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 15

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def prepare_preview_rows(uploaded_files):
    all_rows = []
    for f in uploaded_files:
        parsed_sections = parse_order_bytes(f.getvalue())
        for header_data, items in parsed_sections:
            all_rows.extend(build_row_values(header_data, item) for item in items)
    return all_rows


def convert_uploaded_files(uploaded_files, template_bytes: bytes, selected_rows=None) -> bytes:
    all_rows = prepare_preview_rows(uploaded_files)
    if selected_rows is not None:
        all_rows = selected_rows
    return fill_workbook_from_rows(template_bytes, all_rows)
