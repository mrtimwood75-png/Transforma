import re
import sys
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None


APP_TITLE = "BoConcept Sales Order Converter"
DEFAULT_DELIVERY_TYPE = "premium delivery"
DEFAULT_LOCATION = "bcwh"


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def files_dir() -> Path:
    return app_dir() / "files"


def default_template_path() -> Path:
    folder = files_dir()
    if not folder.exists():
        return Path()
    matches = sorted(folder.glob("*.xlsx"))
    return matches[0] if matches else Path()


def read_text_file(path: str) -> str:
    for enc in ("utf-8-sig", "cp1252", "latin1", "utf-16"):
        try:
            with open(path, "r", encoding=enc, errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    with open(path, "r", errors="ignore") as f:
        return f.read()


def split_rows(text: str):
    rows = []
    for line in text.splitlines():
        rows.append([c.strip() for c in line.rstrip("\n\r").split("\t")])
    return rows


def flatten_non_empty(cells):
    return [str(c).strip() for c in cells if str(c).strip()]


def clean_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def regex_find(text: str, pattern: str, flags=0, default=""):
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else default


def title_case_city(text: str) -> str:
    if not text:
        return ""
    return " ".join(w.capitalize() if not w.isupper() else w for w in text.split())


def normalise_phone(phone: str) -> str:
    return re.sub(r"[^\d+]", "", phone or "")


def extract_sales_order(text: str) -> str:
    m = re.search(r"\b(OS-\d{4,})\b", text, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    return ""


def extract_notes(text: str) -> str:
    lines = [clean_spaces(x) for x in text.splitlines()]
    ignore = {
        "",
        "BoConcept signature",
        "Customer signature",
        "Tax code Amount origin Tax amount",
    }
    candidates = []
    for line in lines:
        if line in ignore:
            continue
        if "Terms & Conditions" in line:
            continue
        if "Payment of any deposit" in line:
            continue
        if "www.bit.ly/BoConceptTerms" in line:
            continue
        if "Trading as BoConcept" in line:
            continue
        if re.search(r"\bSales balance\b|\bBalance due\b|\bPrepayment\b", line, re.I):
            continue
        if any(tag in line.upper() for tag in ["ETA", "DRAWING", "LOANER", "NO ACTION REQUIRED"]):
            candidates.append(line)
    return " | ".join(dict.fromkeys(candidates))


def extract_header_data(text: str, rows):
    first_line = flatten_non_empty(rows[0]) if rows else []
    second_line = flatten_non_empty(rows[1]) if len(rows) > 1 else []

    customer_name = first_line[0] if len(first_line) > 0 else ""
    ship_address = first_line[1] if len(first_line) > 1 else ""

    city = ""
    postcode = ""

    if second_line:
        line2 = " ".join(second_line)
        m = re.search(r"^(.*?)\s+(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+(\d{4})$", line2, re.I)
        if m:
            city = clean_spaces(m.group(1))
            postcode = m.group(2)

    if not city:
        city = regex_find(text, r"\n([A-Z][A-Z\s,]+)\s+(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+\d{4}", re.I)
        city = clean_spaces(city).rstrip(",")

    if not postcode:
        postcode = regex_find(text, r"\b(?:QLD|NSW|VIC|SA|WA|TAS|NT|ACT)\s+(\d{4})\b", re.I)

    email = regex_find(text, r"\bE-Mail\s+([^\s]+@[^\s]+)")
    phone = regex_find(text, r"\bPhone\s+([0-9+ ]{6,})")
    if not phone:
        phone = regex_find(text, r"\bCustomer number\s+([0-9+ ]{6,})")

    return {
        "sales_order_number": extract_sales_order(text),
        "customer_name": customer_name,
        "ship_address": ship_address,
        "ship_zip": postcode,
        "ship_city": title_case_city(city),
        "phone": normalise_phone(phone),
        "email": email,
        "notes": extract_notes(text),
    }


def parse_qty(token: str) -> int:
    token = token.strip().replace(",", ".")
    try:
        return int(round(float(token)))
    except Exception:
        return 1


def normalise_fraction_dimensions(text: str) -> str:
    repl = {
        "½": ".5",
        "¼": ".25",
        "¾": ".75",
        "⅓": ".33",
        "⅔": ".67",
        "⅛": ".125",
        "⅜": ".375",
        "⅝": ".625",
        "⅞": ".875",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    return text


def extract_dimensions(description: str) -> str:
    d = normalise_fraction_dimensions(description or "")

    patterns = [
        r"H\s*\d+(?:[./]\d+)?\s*[xX]\s*W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*D\s*(\d+(?:[./]\d+)?)\s*cm",
        r"W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*D\s*(\d+(?:[./]\d+)?)(?:\s*[xX]\s*H\s*\d+(?:[./]\d+)?)?\s*cm",
        r"W\s*(\d+(?:[./]\d+)?)\s*[xX]\s*L\s*(\d+(?:[./]\d+)?)\s*cm",
        r"\b(\d{2,4}(?:[./]\d+)?)\s*[xX]\s*(\d{2,4}(?:[./]\d+)?)\s*cm\b",
    ]

    for pat in patterns:
        m = re.search(pat, d, re.I)
        if m:
            return f"{m.group(1)} x {m.group(2)}"

    return ""


def extract_full_dimensions(description: str) -> str:
    d = normalise_fraction_dimensions(description or "")

    patterns = [
        r"H\s*\d+(?:[./]\d+)?(?:/\d+(?:[./]\d+)?)?\s*[xX]\s*W\s*\d+(?:[./]\d+)?\s*[xX]\s*D\s*\d+(?:[./]\d+)?\s*cm",
        r"W\s*\d+(?:[./]\d+)?\s*[xX]\s*D\s*\d+(?:[./]\d+)?(?:\s*[xX]\s*H\s*\d+(?:[./]\d+)?)?\s*cm",
        r"W\s*\d+(?:[./]\d+)?\s*[xX]\s*L\s*\d+(?:[./]\d+)?\s*cm",
        r"\b\d{2,4}(?:[./]\d+)?\s*[xX]\s*\d{2,4}(?:[./]\d+)?\s*cm\b",
    ]

    for pat in patterns:
        m = re.search(pat, d, re.I)
        if m:
            return clean_spaces(m.group(0).replace(" ", ""))

    return ""


def parse_items_from_ascii_text(text: str):
    lines = [line.rstrip() for line in text.splitlines()]

    item_start_re = re.compile(
        r"^\s*(\d{4,25})\s+(-?\d+,\d+)\s+(.+?)\s+(?:-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(?:-?\d{1,3}(?:\.\d{3})*,\d{2})\s*$"
    )

    item_start_loose = re.compile(
        r"^\s*(\d{4,25})\s+(-?\d+,\d+)\s+(.+)$"
    )

    stop_re = re.compile(
        r"^\s*(Total for |Tax code\b|gst\d|Sales balance\b|Prepayment\b|Balance due\b)",
        re.I,
    )

    skip_re = re.compile(
        r"^\s*(Article\s+Qty\.?\s+Description|Discount|in pct|per Unit|Price|Amount|Confirmation|Page\s+\d+\s+of\s+\d+)\b",
        re.I,
    )

    footer_re = re.compile(
        r"^\s*(BC Brisbane Pty Ltd|ABN |Tel: |BoConcept.? signature|Customer signature|This order is subject to|Payment of any deposit|You can view BoConcept)",
        re.I,
    )

    items = []
    current = None
    in_items = False

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            continue

        if "Article" in line and "Qty" in line and "Description" in line:
            in_items = True
            continue

        if not in_items:
            continue

        if stop_re.search(line):
            if current:
                current["description"] = clean_spaces(current["description"])
                current["volume"] = extract_dimensions(current["description"])
                current["dimensions"] = extract_full_dimensions(current["description"])
                items.append(current)
                current = None

            if re.match(r"^\s*Tax code\b", line, re.I):
                break

            continue

        if skip_re.search(line) or footer_re.search(line):
            continue

        m = item_start_re.match(line)
        if not m:
            m = item_start_loose.match(line)

        if m:
            if current:
                current["description"] = clean_spaces(current["description"])
                current["volume"] = extract_dimensions(current["description"])
                current["dimensions"] = extract_full_dimensions(current["description"])
                items.append(current)

            sku, qty_raw, desc = m.groups()
            current = {
                "sku": sku,
                "qty": parse_qty(qty_raw),
                "description": clean_spaces(desc),
                "volume": "",
                "dimensions": "",
            }
            continue

        if current:
            if (
                not line.startswith("Total for")
                and "Page " not in line
                and "Sales order " not in line
                and "BC Brisbane Pty Ltd" not in line
            ):
                current["description"] += " " + line

    if current:
        current["description"] = clean_spaces(current["description"])
        current["volume"] = extract_dimensions(current["description"])
        current["dimensions"] = extract_full_dimensions(current["description"])
        items.append(current)

    return items


def parse_order(ascii_path: str):
    text = read_text_file(ascii_path)
    rows = split_rows(text)
    header_data = extract_header_data(text, rows)
    items = parse_items_from_ascii_text(text)
    return header_data, items


def map_headers(ws):
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip().lower()] = cell.column
    return header_map


def set_text(cell, value: str):
    cell.value = "" if value is None else str(value)
    cell.number_format = "@"


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    for col in range(1, max_col + 1):
        s = ws.cell(source_row, col)
        t = ws.cell(target_row, col)
        if s.has_style:
            t._style = s._style.copy()
        if s.font:
            t.font = s.font.copy()
        if s.fill:
            t.fill = s.fill.copy()
        if s.border:
            t.border = s.border.copy()
        if s.alignment:
            t.alignment = s.alignment.copy()
        if s.protection:
            t.protection = s.protection.copy()
        if s.number_format:
            t.number_format = s.number_format


def build_row_values(header_data: dict, item: dict):
    return {
        "sales order number": header_data.get("sales_order_number", ""),
        "delivery type": DEFAULT_DELIVERY_TYPE,
        "customer name": header_data.get("customer_name", ""),
        "sku number": item.get("sku", ""),
        "product description": item.get("description", ""),
        "quantity": item.get("qty", ""),
        "location": DEFAULT_LOCATION,
        "ship address": header_data.get("ship_address", ""),
        "ship zip": header_data.get("ship_zip", ""),
        "ship city": header_data.get("ship_city", ""),
        "phone": header_data.get("phone", ""),
        "weight": "",
        "volume(unit)": item.get("volume", ""),
        "emailaddress": header_data.get("email", ""),
        "notes": header_data.get("notes", ""),
        "dimensions": item.get("dimensions", ""),
    }


def fill_workbook_from_rows(template_path: str, output_path: str, all_rows: list):
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    header_map = map_headers(ws)
    max_col = ws.max_column

    required_headers = [
        "sales order number",
        "delivery type",
        "customer name",
        "sku number",
        "product description",
        "quantity",
        "location",
        "ship address",
        "ship zip",
        "ship city",
        "phone",
        "weight",
        "volume(unit)",
        "emailaddress",
        "notes",
    ]
    for h in required_headers:
        if h not in header_map:
            raise ValueError(f"Template header not found: {h}")

    has_dimensions_col = "dimensions" in header_map

    base_row = 2

    if not all_rows:
        all_rows = [{
            "sales order number": "",
            "delivery type": DEFAULT_DELIVERY_TYPE,
            "customer name": "",
            "sku number": "",
            "product description": "",
            "quantity": "",
            "location": DEFAULT_LOCATION,
            "ship address": "",
            "ship zip": "",
            "ship city": "",
            "phone": "",
            "weight": "",
            "volume(unit)": "",
            "emailaddress": "",
            "notes": "",
            "dimensions": "",
        }]

    if len(all_rows) > 1:
        ws.insert_rows(base_row + 1, amount=len(all_rows) - 1)
        for r in range(base_row + 1, base_row + len(all_rows)):
            copy_row_style(ws, base_row, r, max_col)

    for i, row_values in enumerate(all_rows, start=base_row):
        row_data = dict(row_values)
        if not has_dimensions_col:
            row_data.pop("dimensions", None)

        for header, col in header_map.items():
            if header not in row_data:
                continue

            cell = ws.cell(i, col)
            value = row_data[header]

            if header in {"sku number", "phone", "ship zip", "sales order number"}:
                set_text(cell, value)
            elif header == "quantity":
                cell.value = "" if value == "" else int(value)
            else:
                cell.value = value

    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None:
            ws.column_dimensions[letter].width = 15

    wb.save(output_path)


class PreviewWindow(tk.Toplevel):
    def __init__(self, parent, rows_to_preview):
        super().__init__(parent)
        self.title("Preview items before export")
        self.geometry("1500x680")
        self.resizable(True, True)
        self.result = None
        self.rows_data = [dict(r) for r in rows_to_preview]

        self.columns = [
            "sales order number",
            "customer name",
            "sku number",
            "quantity",
            "product description",
            "dimensions",
            "volume(unit)",
            "ship address",
            "ship city",
            "ship zip",
            "phone",
            "emailaddress",
            "notes",
        ]

        self._build_ui()
        self._load_tree()
        self.grab_set()
        self.transient(parent)

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill="both", expand=True)

        ttk.Label(top, text="All source item rows are shown. Delete only what you do not want exported.").pack(anchor="w", pady=(0, 8))

        tree_frame = ttk.Frame(top)
        tree_frame.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=self.columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        widths = {
            "sales order number": 120,
            "customer name": 140,
            "sku number": 130,
            "quantity": 70,
            "product description": 520,
            "dimensions": 160,
            "volume(unit)": 110,
            "ship address": 220,
            "ship city": 120,
            "ship zip": 80,
            "phone": 110,
            "emailaddress": 180,
            "notes": 220,
        }

        for col in self.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=widths.get(col, 120), anchor="w")

        btns = ttk.Frame(top)
        btns.pack(fill="x", pady=(10, 0))

        ttk.Button(btns, text="Delete selected", command=self.delete_selected).pack(side="left")
        ttk.Button(btns, text="Keep all", command=self.keep_all).pack(side="left", padx=(8, 0))
        ttk.Button(btns, text="Cancel", command=self.cancel).pack(side="right")
        ttk.Button(btns, text="Export", command=self.confirm).pack(side="right", padx=(0, 8))

    def _load_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for idx, row in enumerate(self.rows_data):
            values = [row.get(col, "") for col in self.columns]
            self.tree.insert("", "end", iid=str(idx), values=values)

    def delete_selected(self):
        selected = list(self.tree.selection())
        if not selected:
            return

        selected_indexes = sorted((int(i) for i in selected), reverse=True)
        for idx in selected_indexes:
            if 0 <= idx < len(self.rows_data):
                del self.rows_data[idx]

        self._load_tree()

    def keep_all(self):
        self.result = [dict(r) for r in self.rows_data]
        self.destroy()

    def confirm(self):
        self.result = [dict(r) for r in self.rows_data]
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()


def convert_one(ascii_path: str, template_path: str, output_path: str, preview_parent=None):
    header_data, items = parse_order(ascii_path)
    all_rows = [build_row_values(header_data, item) for item in items]

    if preview_parent is not None:
        preview = PreviewWindow(preview_parent, all_rows)
        preview.wait_window()
        if preview.result is None:
            return False
        all_rows = preview.result

    fill_workbook_from_rows(template_path, output_path, all_rows)
    return True


def convert_many_to_one(ascii_paths: list, template_path: str, output_path: str, preview_parent=None):
    all_rows = []
    for ascii_path in ascii_paths:
        header_data, items = parse_order(ascii_path)
        all_rows.extend(build_row_values(header_data, item) for item in items)

    if preview_parent is not None:
        preview = PreviewWindow(preview_parent, all_rows)
        preview.wait_window()
        if preview.result is None:
            return False
        all_rows = preview.result

    fill_workbook_from_rows(template_path, output_path, all_rows)
    return True


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.withdraw()
        self.title(APP_TITLE)
        self.geometry("780x590")
        self.minsize(780, 590)
        self.configure(bg="white")

        self.input_files = []
        self.output_path = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.mode = tk.StringVar(value="single")
        self.preview_before_export = tk.BooleanVar(value=True)
        self.template_path = default_template_path()

        try:
            style = ttk.Style(self)
            try:
                style.theme_use("vista")
            except Exception:
                try:
                    style.theme_use("clam")
                except Exception:
                    pass

            self._build_ui()
            self.deiconify()
            self.lift()
            self.focus_force()

        except Exception:
            error_text = traceback.format_exc()
            try:
                with open(app_dir() / "startup_error_log.txt", "w", encoding="utf-8") as f:
                    f.write(error_text)
            except Exception:
                pass
            self.destroy()
            raise

    def _build_ui(self):
        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        self._build_header(outer)

        ttk.Label(outer, text="ASCII sales order file(s)").pack(anchor="w", pady=(10, 4))
        files_frame = ttk.Frame(outer)
        files_frame.pack(fill="x")

        self.files_list = tk.Listbox(files_frame, height=10, selectmode=tk.EXTENDED)
        self.files_list.pack(side="left", fill="both", expand=True)

        btns = ttk.Frame(files_frame)
        btns.pack(side="left", padx=(10, 0), fill="y")
        ttk.Button(btns, text="Add file(s)", command=self.pick_input_files).pack(fill="x", pady=2)
        ttk.Button(btns, text="Clear", command=self.clear_input_files).pack(fill="x", pady=2)

        mode_frame = ttk.LabelFrame(outer, text="Output mode", padding=10)
        mode_frame.pack(fill="x", pady=(12, 8))

        ttk.Radiobutton(
            mode_frame,
            text="Single file to single workbook",
            variable=self.mode,
            value="single",
            command=self.toggle_output_mode,
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="Multiple files to one workbook",
            variable=self.mode,
            value="many_to_one",
            command=self.toggle_output_mode,
        ).pack(anchor="w")
        ttk.Radiobutton(
            mode_frame,
            text="Multiple files to separate workbooks",
            variable=self.mode,
            value="batch",
            command=self.toggle_output_mode,
        ).pack(anchor="w")

        ttk.Checkbutton(
            outer,
            text="Show preview window before export",
            variable=self.preview_before_export,
        ).pack(anchor="w", pady=(2, 8))

        template_text = f"Template: {self.template_path}" if self.template_path else "Template not found in ./files"
        self.template_label = ttk.Label(outer, text=template_text)
        self.template_label.pack(anchor="w", pady=(2, 8))

        self.output_label = ttk.Label(outer, text="Output workbook")
        self.output_label.pack(anchor="w", pady=(8, 4))
        self.output_frame = self._path_row(outer, self.output_path, self.pick_output_file)

        self.batch_output_label = ttk.Label(outer, text="Output folder")
        self.batch_output_frame = self._path_row(outer, self.output_folder, self.pick_output_folder)
        self.batch_output_label.pack_forget()
        self.batch_output_frame.pack_forget()

        ttk.Button(outer, text="Convert", command=self.run_conversion).pack(anchor="e", pady=(20, 0))

    def _build_header(self, parent):
        header = ttk.Frame(parent)
        header.pack(fill="x")

        logo_loaded = False
        logo_paths = [
            files_dir() / "boconcept_logo.png",
            files_dir() / "boconcept_logo.jpg",
            app_dir() / "boconcept_logo.png",
            app_dir() / "boconcept_logo.jpg",
            app_dir() / "logo.png",
            app_dir() / "logo.jpg",
        ]

        if Image and ImageTk:
            for p in logo_paths:
                if p.exists():
                    try:
                        img = Image.open(p)
                        img.thumbnail((220, 80))
                        self.logo_img = ImageTk.PhotoImage(img)
                        tk.Label(header, image=self.logo_img, bg="white").pack(anchor="w")
                        logo_loaded = True
                        break
                    except Exception:
                        pass

        if not logo_loaded:
            tk.Label(
                header,
                text="BoConcept",
                font=("Arial", 24, "bold"),
                fg="black",
                bg="white",
            ).pack(anchor="w")

        ttk.Label(
            header,
            text="Sales Order ASCII to Excel Converter",
            font=("Arial", 11),
        ).pack(anchor="w", pady=(6, 0))

    def _path_row(self, parent, var, command):
        frame = ttk.Frame(parent)
        frame.pack(fill="x")
        entry = ttk.Entry(frame, textvariable=var)
        entry.pack(side="left", fill="x", expand=True)
        ttk.Button(frame, text="Browse", command=command).pack(side="left", padx=(8, 0))
        return frame

    def pick_input_files(self):
        paths = filedialog.askopenfilenames(
            title="Select ASCII sales order file(s)",
            filetypes=[("Text files", "*.txt *.asc *.csv"), ("All files", "*.*")],
        )
        if paths:
            self.input_files = list(paths)
            self.files_list.delete(0, tk.END)
            for p in self.input_files:
                self.files_list.insert(tk.END, p)

    def clear_input_files(self):
        self.input_files = []
        self.files_list.delete(0, tk.END)

    def pick_output_file(self):
        path = filedialog.asksaveasfilename(
            title="Save output workbook as",
            defaultextension=".xlsx",
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if path:
            self.output_path.set(path)

    def pick_output_folder(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.output_folder.set(path)

    def toggle_output_mode(self):
        mode = self.mode.get()
        if mode == "batch":
            self.output_label.pack_forget()
            self.output_frame.pack_forget()
            self.batch_output_label.pack(anchor="w", pady=(8, 4))
            self.batch_output_frame.pack(fill="x")
        else:
            self.batch_output_label.pack_forget()
            self.batch_output_frame.pack_forget()
            self.output_label.pack(anchor="w", pady=(8, 4))
            self.output_frame.pack(fill="x")

    def run_conversion(self):
        try:
            if not self.template_path or not Path(self.template_path).exists():
                raise ValueError("Template not found. Put the Excel template in a folder called 'files' next to the app.")

            if not self.input_files:
                raise ValueError("Select at least one ASCII input file.")

            mode = self.mode.get()
            use_preview = self.preview_before_export.get()

            if mode == "single":
                if len(self.input_files) != 1:
                    raise ValueError("Single mode requires exactly one input file.")
                if not self.output_path.get():
                    raise ValueError("Select the output workbook location.")
                ok = convert_one(
                    self.input_files[0],
                    str(self.template_path),
                    self.output_path.get(),
                    preview_parent=self if use_preview else None,
                )
                if ok:
                    messagebox.showinfo("Done", f"Created:\n\n{self.output_path.get()}")

            elif mode == "many_to_one":
                if not self.output_path.get():
                    raise ValueError("Select the output workbook location.")
                ok = convert_many_to_one(
                    self.input_files,
                    str(self.template_path),
                    self.output_path.get(),
                    preview_parent=self if use_preview else None,
                )
                if ok:
                    messagebox.showinfo("Done", f"Created combined workbook:\n\n{self.output_path.get()}")

            elif mode == "batch":
                if not self.output_folder.get():
                    raise ValueError("Select an output folder.")
                out_dir = Path(self.output_folder.get())
                out_dir.mkdir(parents=True, exist_ok=True)

                results = []
                for inp in self.input_files:
                    stem = Path(inp).stem
                    out_file = out_dir / f"{stem}_converted.xlsx"
                    ok = convert_one(
                        inp,
                        str(self.template_path),
                        str(out_file),
                        preview_parent=self if use_preview else None,
                    )
                    if ok:
                        results.append(str(out_file))

                if results:
                    messagebox.showinfo("Done", "Created:\n\n" + "\n".join(results))

        except Exception:
            error_text = traceback.format_exc()
            try:
                with open(app_dir() / "runtime_error_log.txt", "w", encoding="utf-8") as f:
                    f.write(error_text)
            except Exception:
                pass
            messagebox.showerror("Error", error_text)


def run_cli():
    tmpl = default_template_path()

    if len(sys.argv) >= 3:
        if not tmpl or not tmpl.exists():
            raise FileNotFoundError("Template not found in ./files")

        output_path = sys.argv[-1]
        ascii_paths = sys.argv[1:-1]

        all_rows = []
        for ascii_path in ascii_paths:
            header_data, items = parse_order(ascii_path)
            all_rows.extend(build_row_values(header_data, item) for item in items)

        fill_workbook_from_rows(str(tmpl), output_path, all_rows)
        print(f"Created: {output_path}")
    else:
        app = App()
        app.mainloop()


def main():
    try:
        if len(sys.argv) >= 3:
            run_cli()
        else:
            app = App()
            app.mainloop()
    except Exception:
        error_text = traceback.format_exc()
        try:
            with open(app_dir() / "fatal_error_log.txt", "w", encoding="utf-8") as f:
                f.write(error_text)
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()
